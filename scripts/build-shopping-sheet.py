#!/usr/bin/env python3
"""Rebuild out/shopping-sites.xlsx from out/shopping-sites.csv.

The CSV is the working copy; the Google Sheet named in CLAUDE.md is canonical.
Run this after editing the CSV so the formatted workbook stays in sync.
"""
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

FONT = "Arial"
FILLS = {"OK": "E8F5E9", "Caution": "FFF4E0", "Avoid": "FDE7E7"}
ORDER = ["Bags & Purses", "Clothing", "Footwear", "Accessories", "Beauty",
         "Resale", "Aggregator", "Marketplace", "Home", "Supplements",
         "Food", "Travel", "Utility"]
WIDTHS = dict(zip("ABCDEFG", [24, 14, 26, 34, 10, 42, 70]))


def main():
    rows = list(csv.reader(open("out/shopping-sites.csv")))
    hdr, data = rows[0], rows[1:]
    data.sort(key=lambda r: (ORDER.index(r[1]), r[0].lower()))

    wb = Workbook()
    ws = wb.active
    ws.title = "Websites"
    ws.append(hdr)
    for r in data:
        ws.append(r)

    thin = Side(style="thin", color="D0D0D0")
    for c in ws[1]:
        c.font = Font(name=FONT, bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill("solid", fgColor="3A3A3A")
        c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        trust = row[4].value
        for c in row:
            c.font = Font(name=FONT, size=10)
            c.alignment = Alignment(vertical="top", wrap_text=c.column in (6, 7))
            c.border = Border(bottom=thin)
            c.fill = PatternFill("solid", fgColor=FILLS[trust])
        url = row[3]
        url.hyperlink = url.value
        url.font = Font(name=FONT, size=10, color="0563C1", underline="single")
        row[4].alignment = Alignment(horizontal="center", vertical="top")
        row[4].font = Font(name=FONT, size=10, bold=(trust != "OK"))

    for col, width in WIDTHS.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{ws.max_row}"

    add_summary(wb, ws, data)
    wb.save("out/shopping-sites.xlsx")
    print(f"wrote out/shopping-sites.xlsx — {ws.max_row - 1} rows")


def add_summary(wb, ws, data):
    s = wb.create_sheet("Summary")
    last = ws.max_row
    cats = list(dict.fromkeys(r[1] for r in data))
    s["A1"], s["B1"] = "Category", "Sites"
    for n, cat in enumerate(cats, start=2):
        s[f"A{n}"] = cat
        s[f"B{n}"] = f'=COUNTIF(Websites!$B$2:$B${last},A{n})'
    end = len(cats) + 1
    s[f"A{end + 1}"], s[f"B{end + 1}"] = "Total", f"=SUM(B2:B{end})"
    s[f"A{end + 3}"], s[f"B{end + 3}"] = "Trust", "Sites"
    for j, t in enumerate(["OK", "Caution", "Avoid"], start=end + 4):
        s[f"A{j}"] = t
        s[f"B{j}"] = f'=COUNTIF(Websites!$E$2:$E${last},A{j})'
    note = ("Source: browser history 14-22 Aug 2026. Trust ratings are a judgement call on "
            "authenticity, privacy and seller legitimacy - not a formal security scan.")
    s[f"A{end + 8}"] = note
    for row in s.iter_rows(min_row=1, max_row=s.max_row):
        for c in row:
            c.font = Font(name=FONT, size=10)
    for cell in ("A1", "B1", f"A{end + 1}", f"B{end + 1}", f"A{end + 3}", f"B{end + 3}"):
        s[cell].font = Font(name=FONT, size=10, bold=True)
    s[f"A{end + 8}"].font = Font(name=FONT, size=9, italic=True, color="777777")
    s.column_dimensions["A"].width = 26
    s.column_dimensions["B"].width = 10


if __name__ == "__main__":
    main()
